import discord
from discord.ext import commands
from discord import app_commands
import aiohttp
import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import re
import locale
import requests
import time

# =====================
# CONFIGURATION
# =====================
TOKEN = ""
EXCEL_FILE = "anime_zestawienie.xlsx"
ANILIST_API_URL = "https://graphql.anilist.co"

intents = discord.Intents.default()
intents.message_content = True
bot = commands.Bot(command_prefix="!", intents=intents)

# Set locale (for Polish months)
try:
    locale.setlocale(locale.LC_TIME, 'pl_PL.UTF-8')
except locale.Error:
    locale.setlocale(locale.LC_TIME, '')

# =====================
# HELPERS
# =====================

def extract_titles(text):
    """Extract AniList and MyAnimeList titles/IDs from text."""
    titles = set()
    patterns = [
        (r"\[([^\]]+)\]\(https:\/\/anilist\.co\/anime\/(\d+)\)", lambda m: (m.group(2), None)),
        (r"https:\/\/anilist\.co\/anime\/(\d+)", lambda m: (m.group(1), None)),
        (r"\[([^\]]+)\]\(https:\/\/myanimelist\.net\/anime\/(\d+)\)", lambda m: (None, m.group(1).strip())),
        (r"https:\/\/myanimelist\.net\/anime\/\d+\/([a-zA-Z0-9\-_]+)", lambda m: (None, m.group(1).replace("_", " ").replace("-", " ").strip()))
    ]
    for pattern, extractor in patterns:
        for match in re.finditer(pattern, text):
            titles.add(extractor(match))
    return list(titles)


def get_anilist_titles_batch(ids: list[int]) -> dict:
    """
    Fetch multiple AniList titles at once with automatic exponential backoff on 429.
    Returns a dict {id: romaji_title}.
    """
    if not ids:
        return {}

    query_parts = []
    for idx, anime_id in enumerate(ids):
        query_parts.append(f"anime{idx}: Media(id: {anime_id}) {{ title {{ romaji }} }}")

    query = "query {" + " ".join(query_parts) + "}"
    result = {}

    wait_time = 1  # initial wait for backoff
    while True:
        try:
            response = requests.post(ANILIST_API_URL, json={"query": query})
            if response.status_code == 429:
                print(f"⚠️ Rate limited. Waiting {wait_time}s before retry...")
                time.sleep(wait_time)
                wait_time = min(wait_time * 2, 30)  # exponential backoff up to 30s
                continue
            response.raise_for_status()
            data = response.json().get("data", {})

            for idx, anime_id in enumerate(ids):
                key = f"anime{idx}"
                romaji = data.get(key, {}).get("title", {}).get("romaji")
                if romaji:
                    result[anime_id] = romaji
                else:
                    result[anime_id] = f"AniList ID {anime_id}"  # fallback
            break  # exit loop if successful

        except Exception as e:
            print(f"❌ Error fetching batch: {e}")
            # Assign fallback for all remaining IDs
            for anime_id in ids:
                if anime_id not in result:
                    result[anime_id] = f"AniList ID {anime_id}"
            break

    return result


def search_anilist_by_title(title: str) -> str:
    """Search AniList by title string and return the best match Romaji title. Fallback if no match."""
    query = """
    query ($search: String) {
      Media(search: $search, type: ANIME) {
        title { romaji }
      }
    }
    """
    variables = {"search": title}

    try:
        response = requests.post(ANILIST_API_URL, json={"query": query, "variables": variables})
        if response.status_code == 429:
            print("⚠️ Rate limited during search. Waiting 2s...")
            time.sleep(2)
            response = requests.post(ANILIST_API_URL, json={"query": query, "variables": variables})
        response.raise_for_status()
        data = response.json()
        media = data.get("data", {}).get("Media")
        if not media:
            return title  # fallback

        return media.get("title", {}).get("romaji") or title
    except Exception as e:
        print(f"❌ AniList search error for '{title}': {e}")
        return title


def log_error(value):
    """Log unknown entries to file."""
    with open("bledy.txt", "a", encoding="utf-8") as f:
        f.write(f"Nie rozpoznano: {value}\n")


def save_to_excel(user, anime, date):
    """Save anime record to Excel (avoid duplicates)."""
    month = date.strftime("%B").capitalize()
    date_str = date.strftime("%Y-%m-%d")

    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Uczestnik", "Anime", "Miesiąc", "Data"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    if any(
        row[0] == user and row[1] == anime and row[3] == date_str
        for row in ws.iter_rows(min_row=2, values_only=True)
    ):
        return False

    ws.append([user, anime, month, date_str])
    wb.save(EXCEL_FILE)
    return True

# =====================
# COMMANDS
# =====================

@bot.event
async def on_ready():
    await bot.tree.sync()
    print(f"✅ Logged in as {bot.user}")


@bot.tree.command(name="analizuj_watek", description="Analizuje podany wątek i zapisuje anime do Excela")
@app_commands.describe(link="Link do wątku Discord")
async def analizuj_watek(interaction: discord.Interaction, link: str):
    await interaction.response.defer(thinking=True)

    try:
        match = re.search(r"channels/\d+/(\d+)", link)
        if not match:
            await interaction.followup.send("❌ Nieprawidłowy link do wątku.")
            return

        thread_id = int(match.group(1))
        thread = await bot.fetch_channel(thread_id)

        if not isinstance(thread, discord.Thread):
            await interaction.followup.send("❌ Podany link nie prowadzi do wątku.")
            return

        messages = [msg async for msg in thread.history(limit=None)]
        if not messages:
            await interaction.followup.send("❌ Brak wiadomości do analizy.")
            return

        author = messages[-1].author
        await interaction.followup.send(f"🔍 Analizuję wiadomości użytkownika: **{author.display_name}**...")

        # Gather all AniList IDs for batch fetching
        all_ids = []
        msg_titles_mapping = {}  # map message to list of (anime_id, fallback_title)
        for msg in reversed(messages):
            if msg.author != author:
                continue
            titles_list = extract_titles(msg.content)
            msg_titles_mapping[msg] = titles_list
            for anime_id, _ in titles_list:
                if anime_id:
                    all_ids.append(int(anime_id))

        # Fetch all titles in batch
        titles_dict = get_anilist_titles_batch(all_ids)

        # Save to Excel
        saved, skipped = 0, 0
        for msg, titles_list in msg_titles_mapping.items():
            for anime_id, fallback_title in titles_list:
                if anime_id:
                    title = titles_dict.get(int(anime_id))
                else:
                    title = search_anilist_by_title(fallback_title)

                if not title:
                    log_error(f"AniList ID {anime_id}" if anime_id else f"MAL fallback: {fallback_title}")
                    title = f"AniList ID {anime_id}" if anime_id else fallback_title

                title = title.strip().rstrip("/")
                if save_to_excel(msg.author.display_name, title, msg.created_at):
                    saved += 1
                else:
                    skipped += 1

        await interaction.followup.send(f"✅ Gotowe! Dodano **{saved}** nowych tytułów. Pominięto **{skipped}** (duplikaty).")

    except Exception as e:
        print(f"[ERROR] {e}")
        await interaction.followup.send("⚠️ Wystąpił błąd podczas analizy. Upewnij się, że link do wątku jest poprawny.")


@bot.tree.command(name="react_post", description="Reaguje na wskazany post emoji")
@app_commands.describe(link="Link do wiadomości", emoji="Emoji do reakcji (np. ❤️ lub 👍)")
async def react_post(interaction: discord.Interaction, link: str, emoji: str):
    await interaction.response.defer(thinking=True)
    try:
        match = re.search(r"channels/\d+/(\d+)/(\d+)", link)
        if not match:
            await interaction.followup.send("❌ Nieprawidłowy link do wiadomości.")
            return

        channel_id, message_id = map(int, match.groups())
        channel = await bot.fetch_channel(channel_id)
        message = await channel.fetch_message(message_id)

        await message.add_reaction(emoji)
        await interaction.followup.send(f"✅ Dodano reakcję {emoji} do wiadomości: {message.jump_url}")
    except Exception as e:
        print(f"[Reaction error] {e}")
        await interaction.followup.send("⚠️ Nie udało się dodać reakcji. Sprawdź link i emoji.")


# =====================
# RUN
# =====================
bot.run(TOKEN)